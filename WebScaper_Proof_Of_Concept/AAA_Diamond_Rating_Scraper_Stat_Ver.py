import openpyxl
import warnings
import string
from selenium import webdriver
from time import sleep
from fuzzywuzzy import fuzz

# suppress static warnings
warnings.filterwarnings("ignore")


def get_hotel_list(file_path):
    """
    Get List of Hotels from  Excel List

    :param file_path: input file
    :return: list of hotels with key information
    """

    # Create file object then create reference to first Sheet in Excel file
    wb = openpyxl.load_workbook(filename=file_path)
    sheet = wb.worksheets[0]  # 0 is index of sheet in excel file

    # column range in which you want to pull
    excel_col_range = ['A', 'B', 'C', 'D', 'E', 'F', 'G']

    # create list to store hotel data
    hotel_list = []

    # Start at range 1 since there are headers in excel file template
    for row in range(1, sheet.max_row + 1):
        # list of all hotel attributes
        individual_hotel = []

        # from row in file iterate all data in excel_col_range
        for col in excel_col_range:
            individual_hotel.append(sheet[col + str(row)].value)

        # store single hotel in hotel_list
        hotel_list.append(individual_hotel)

    return hotel_list


def get_diamond_rating(html_source):
    """
    Get the diamond rating from the html code

    :param html_source: html javascript source w/ information about diamond rating
    :return: diamond rating (black or red)
    """

    # conventional format of diamond rating string
    red_heading = 'rs_icon rs_diamond_red rs_diamond_red--'
    black_heading = 'rs_icon rs_diamond_black rs_diamond_black--'

    if red_heading + '5' in html_source:
        return 'Red Diamond Award: 5'
    elif red_heading + '4' in html_source:
        return 'Red Diamond Award: 4'
    elif red_heading + '3' in html_source:
        return 'Red Diamond Award: 3'
    elif red_heading + '2' in html_source:
        return 'Red Diamond Award: 2'
    elif red_heading + '1' in html_source:
        return 'Red Diamond Award: 1'
    elif black_heading + '5' in html_source:
        return 'Black Diamond Award: 5'
    elif black_heading + '4' in html_source:
        return 'Black Diamond Award: 4'
    elif black_heading + '3' in html_source:
        return 'Black Diamond Award: 3'
    elif black_heading + '2' in html_source:
        return 'Black Diamond Award: 2'
    elif black_heading + '1' in html_source:
        return 'Black Diamond Award: 1'
    else:
        return 'No Diamond Award Found'


def view_media_content(browser):
    """
    Content such as user rating and other etc information about hotels can be viewed

    information such as: Street Address, User Ratings, and Hotel Amenities
    :param browser: web driver object reference
    :return: sting text of information about hotel
    """

    hotel_content = browser.find_element_by_class_name('rs_media_content')

    if hotel_content:
        print hotel_content.text


def main():

    # Select the Web Driver Chrome(), Firefox(), PhantomJS, Etc.
    browser = webdriver.Chrome()

    reset_counter = 0

    # Create excel workbook object to store results
    workbook = openpyxl.Workbook()
    sheet1 = workbook.active  # reference to sheet in workbook

    # append headers to new sheet
    sheet1.append(['LocationID', 'CustomerName', 'City', 'State', 'Country', 'Primary Category', 'Diamond Rating'])

    # file to list of hotels with AAA city id's
    file_path = 'List_of_All_Locations.xlsx'

    # list of hotels taken from excel file via file_path
    hotel_list = get_hotel_list(file_path)

    # traverse through hotel_list and collect AAA Diamond Rating Data
    for single_hotel in hotel_list[1:]:

        location_id = single_hotel[0]
        hotel_name = single_hotel[1]
        city = single_hotel[2]
        state = single_hotel[3]
        country = single_hotel[4]
        primary_category = single_hotel[5]
        city_id = single_hotel[6]

        hotel_name_temp = hotel_name

        if '&' in hotel_name:
            hotel_name_temp = hotel_name.replace(hotel_name[hotel_name.index('&')], '&amp;')

        hotel_name_list = hotel_name_temp.split()

        if '/' in hotel_name_list:
            hotel_name_list.remove('/')

        if '\\' in hotel_name_list:
            hotel_name_list.remove('\\')

        if '-' in hotel_name_list:
            hotel_name_list.remove('-')

        if '!' in hotel_name_list:
            hotel_name_list.remove('!')

        if 'hotel' in hotel_name_list:
            hotel_name_list.remove('hotel')

        if 'Hotel' in hotel_name_list:
            hotel_name_list.remove('Hotel')

        hotel_name_temp = ' '.join(hotel_name_list)

        # URL of AAA website in which we are trying to web scrape
        url = 'http://secure.rezserver.com/hotels/results_v2/?refid=5734&varid=1b2b3b4b&rooms=1&adults=2' \
              '&hotel_name=' + hotel_name_temp.replace(" ", "%20") + '&city_id=' + str(city_id)

        # encoding rule
        url.encode('utf-8').strip()

        # execute web driver to get url
        browser.get(url)

        # Get all javascript modified html source

        try:
            sleep(5)
            html_source = browser.execute_script("return "
                                                 "document.getElementsByClassName('rs_media_content')[0].innerHTML")

            html_source_hotel = browser.execute_script("return "
                                                       "document.getElementsByClassName('rs_media_content')[0]"
                                                       ".getElementsByClassName('rs_hotel_name rs_a_style')[0]"
                                                       ".innerHTML")

            str_html_source = html_source.encode('ascii', 'ignore')

            print hotel_name, 'VS', html_source_hotel

            p1 = fuzz.ratio(hotel_name, html_source_hotel)
            p2 = fuzz.partial_ratio(hotel_name, html_source_hotel)
            p3 = fuzz.token_sort_ratio(hotel_name, html_source_hotel)
            p4 = fuzz.token_set_ratio(hotel_name, html_source_hotel)
            print 'ratio', p1
            print 'partial_ratio', p2
            print 'token_sort_ratio', p3
            print 'token_set_ratio', p4

            print hotel_name_temp, 'VS', html_source_hotel

            t1 = fuzz.ratio(hotel_name_temp, html_source_hotel)
            t2 = fuzz.partial_ratio(hotel_name_temp, html_source_hotel)
            t3 = fuzz.token_sort_ratio(hotel_name_temp, html_source_hotel)
            t4 = fuzz.token_set_ratio(hotel_name_temp, html_source_hotel)
            print 'ratio', t1
            print 'partial_ratio', t2
            print 'token_sort_ratio', t3
            print 'token_set_ratio', t4

            # if hotel_name found in html_source then hotel was found in AAA Website

            if p1 > 80 or p2 > 80 or p3 > 80 or p4 > 80 or t1 > 80 or t2 > 80 or t3 > 80 or t4 > 80:
                print 'Property Found:', hotel_name

                # view_media_content(browser)

                # get diamond rating if it exist
                diamond_rating = get_diamond_rating(str_html_source)
                print diamond_rating

                # append results to sheet1 object where all results will be found
                sheet1.append([location_id, hotel_name, city, state, country, primary_category, diamond_rating])

        except:
            print 'Hotel Name:', hotel_name, "Error: rs_media_content not found"

        print '----------------------------------------------------------------------------------'

        reset_counter += 1

        # release memory reference to selenium every 300 executions to prevent crashing
        if reset_counter == 300:
            browser = webdriver.Chrome()
            reset_counter = 0

    # Save results of web scrape to file with name from save_file_name
    save_file_name = "AAA_Web_Scraping_Results.xlsx"
    workbook.save(save_file_name)

    print "Web Scraping Complete"

if __name__ == '__main__':
    main()
